import csv
import openpyxl

def get_device(d):
    wb = openpyxl.load_workbook('excel.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    with open(d+'.csv', 'w') as csvfile:
        filewriter = csv.writer(csvfile, delimiter=',',quotechar='|', quoting=csv.QUOTE_MINIMAL)
        for i in range(4,938,1):
            a_device = sheet.cell(row=i, column=11).value
            a_port = sheet.cell(row=i, column=9).value
            atrunk_port = sheet.cell(row=i, column=13).value
            ztrunk_port = sheet.cell(row=i, column=26).value
            z_device = sheet.cell(row=i, column=28).value
            z_port = sheet.cell(row=i, column=30).value
            if a_device == d:
                filewriter.writerow([a_device,a_port,z_device,z_port])
        for i in range(4,938,1):
            a_device = sheet.cell(row=i, column=11).value
            a_port = sheet.cell(row=i, column=9).value
            atrunk_port = sheet.cell(row=i, column=13).value
            ztrunk_port = sheet.cell(row=i, column=26).value
            z_device = sheet.cell(row=i, column=28).value
            z_port = sheet.cell(row=i, column=30).value
            if a_device == d:
                filewriter.writerow([a_device,atrunk_port,z_device,ztrunk_port])

def main():
    wb = openpyxl.load_workbook('excel.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    rock_racks = []
    for i in range(4,938,1):
        if sheet.cell(row=i, column=11).value in rock_racks or not sheet.cell(row=i, column=11).value:
            continue
        else:
            rock_racks.append(sheet.cell(row=i, column=11).value)
    for i in rock_racks:
        create_csv = get_device(i)

if __name__ == '__main__':
    main()
