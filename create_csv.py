import csv
import openpyxl
from prettytable import PrettyTable
import argparse
import logging 

parser = argparse.ArgumentParser(description='python script for generating CSV cutsheet')
parser.add_argument('--fromrow', help='Enter first row to include')
parser.add_argument('--torow', help='Enter last row to include')
parser.add_argument('--columns', help='Enter which columns to include', nargs='+')
parser.add_argument('--input',help='Enter excel file to read')
parser.add_argument('--output', help='enter csv filename')
parser.add_argument('--sheet',help='Enter sheet name ')
parser.add_argument('--router', help='Enter router name')
args=parser.parse_args()

def main():
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    logger.info('Opening excel file')
    wb = openpyxl.load_workbook(args.input)
    sheet = wb.get_sheet_by_name(args.sheet)
    t = PrettyTable()
    t.field_names = ['hostname','interface_name','speed','cage','rack','patch_panel','ports']
    with open(args.output, 'w') as csvfile:
        filewriter = csv.writer(csvfile, delimiter=',',quotechar='|', quoting=csv.QUOTE_MINIMAL)
        for i in range(int(args.fromrow),int(args.torow),1):
            c1 = args.router
            c2 =  sheet.cell(row=i, column=int(args.columns[0])).value
            c3 =  sheet.cell(row=i, column=int(args.columns[1])).value
            c4 =  sheet.cell(row=i, column=int(args.columns[2])).value
            c5 =  sheet.cell(row=i, column=int(args.columns[3])).value
            c6 =  sheet.cell(row=i, column=int(args.columns[4])).value
            c7 =  sheet.cell(row=i, column=int(args.columns[5])).value
            filewriter.writerow([c1,c2,c3,c4,c5,c6,c7])
    logger.info('Finish writing rows to file')
if __name__ == '__main__':
    main()
