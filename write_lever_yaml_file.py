import openpyxl
import yaml
import re
import argparse
import logging
from tqdm import tqdm

parser = argparse.ArgumentParser(description='python script for generating yaml files')
parser.add_argument('--input',help='Enter input file')
parser.add_argument('--output',help='Enter output file')
parser.add_argument('--sheet',help='Enter sheet name')
args=parser.parse_args()

#setting up logging
logging.basicConfig(format='%(asctime)s - %(message)s', level=logging.INFO)

def position_check(s):
    return (int(s.split('-')[0][3:])+1)//2

def stripe_check(s):
    return s[-1:]

def get_interfaces(idx,sheet):
    #Grab interfaces
    return [sheet.cell(row=i,column=2).value for i in range(idx,idx+4,1)]

def validate_file(sheet):
    lines_read = 0
    lines_ignored = 0
    list_of_lbes = []
    lbe_pattern = re.compile(r"\w+\d+\-br\-\w+\-j\d+\-r\d+")
    my_dict = {}
    for i in tqdm(range(1,sheet.max_row+1,1)):
        lbe_router = sheet.cell(row=i,column=1).value
        match_lbe = lbe_pattern.match(lbe_router)
        if not match_lbe:
            lines_ignored += 1
            continue
        else:
            lines_read += 1
            if lbe_router in list_of_lbes:
                continue
            else:
                list_of_lbes.append(lbe_router)
    logging.info('{} lines have been read'.format(lines_read))
    logging.info('{} lines have been ignored'.format(lines_ignored))
    logging.info('{} LBEs have been detected'.format(len(list_of_lbes)))
    return sorted(list_of_lbes)

def open_file(file):
    #Open excel file
    try:
        wb = openpyxl.load_workbook(file)
    except:
        print('Could not open {}'.format(file))

def write_yaml_file(data,output_file):
    #Write to file
    try:
        with open(output_file, 'w') as f:
            yaml.dump(data, f, default_flow_style=False, indent=4)
    except:
        print('Something went wrong, {} could not be opened for writing'.format(output_file))

def dummy_data(idx,sheet,my_dict):
    for i in range(idx,idx+15,15):
        logging.info('Reading rows from {} to {}'.format(idx,idx+15))

        lbe_router1 = '{}'.format(sheet.cell(row=i,column=1).value)
        lbe_router2 = '{}{}'.format(sheet.cell(row=i,column=1).value,'_2')
        lbe_router3 = '{}{}'.format(sheet.cell(row=i,column=1).value,'_3')
        lbe_router4 = '{}{}'.format(sheet.cell(row=i,column=1).value,'_4')
        my_dict[lbe_router1] = {}
        my_dict[lbe_router1]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(sheet.cell(row=i,column=2).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+2,column=2).value,sheet.cell(row=i+3,column=2).value)
        my_dict[lbe_router1]['position'] = position_check()
        my_dict[lbe_router1]['stripe'] = stripe_check()
        my_dict[lbe_router2] = {}
        my_dict[lbe_router2]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(sheet.cell(row=i+4,column=2).value,sheet.cell(row=i+5,column=2).value,sheet.cell(row=i+6,column=2).value,sheet.cell(row=i+7,column=2).value)
        my_dict[lbe_router3] = {}
        my_dict[lbe_router3]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(sheet.cell(row=i+8,column=2).value,sheet.cell(row=i+9,column=2).value,sheet.cell(row=i+10,column=2).value,sheet.cell(row=i+11,column=2).value)
        my_dict[lbe_router4] = {}
        my_dict[lbe_router4]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(sheet.cell(row=i+12,column=2).value,sheet.cell(row=i+13,column=2).value,sheet.cell(row=i+14,column=2).value,sheet.cell(row=i+15,column=2).value)
        logging.info('{} has been created'.format(lbe_router1))
        logging.info('{} has been created'.format(lbe_router2))
        logging.info('{} has been created'.format(lbe_router3))
        logging.info('{} has been created'.format(lbe_router4))

def main():
    open_file(args.input)
    wb = openpyxl.load_workbook(args.input)
    sheet = wb[args.sheet]
    lbe_routers_list = validate_file(sheet)
    counter = 0
    my_dict = {}
    id = 1
    for i in range(1,sheet.max_row+1,4):
        logging.info('Current index is {}'.format(i))
        int1 = sheet.cell(row=i,column=2).value
        int2 = sheet.cell(row=i+1,column=2).value
        int3 = sheet.cell(row=i+2,column=2).value
        int4 = sheet.cell(row=i+3,column=2).value
        position = position_check(sheet.cell(row=i,column=9).value)
        stripe = stripe_check(sheet.cell(row=i,column=11).value)
        if sheet.cell(row=i,column=1).value not in my_dict:
            lbe_router = sheet.cell(row=i,column=1).value
            my_dict[lbe_router] = {}
            my_dict[lbe_router]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(int1,int2,int3,int4)
            my_dict[lbe_router]['position'] = position
            my_dict[lbe_router]['stripe'] = stripe
            counter += 1
            id += 1
            continue
        if id == 2:
            lbe_router = '{}{}{}'.format(sheet.cell(row=i,column=1).value,'_',id)
            my_dict[lbe_router] = {}
            my_dict[lbe_router]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(int1,int2,int3,int4)
            my_dict[lbe_router]['position'] = position
            my_dict[lbe_router]['stripe'] = stripe
            id += 1
            counter += 1
            continue
        if id == 3:
            lbe_router = '{}{}{}'.format(sheet.cell(row=i,column=1).value,'_',id)
            my_dict[lbe_router] = {}
            my_dict[lbe_router]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(int1,int2,int3,int4)
            my_dict[lbe_router]['position'] = position
            my_dict[lbe_router]['stripe'] = stripe
            id += 1
            counter += 1
            continue
        if id == 4:
            lbe_router = '{}{}{}'.format(sheet.cell(row=i,column=1).value,'_',id)
            my_dict[lbe_router] = {}
            my_dict[lbe_router]['remote_edge_interfaces'] = '[{},{},{},{}]'.format(int1,int2,int3,int4)
            my_dict[lbe_router]['position'] = position
            my_dict[lbe_router]['stripe'] = stripe
            counter += 1
            id = 1


    logging.info('{} entries have been created'.format(counter))
    write_file = write_yaml_file(my_dict,args.output)


if __name__ == '__main__':
    main()
